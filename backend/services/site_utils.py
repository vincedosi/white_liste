"""Pure helpers for site/domain normalisation — no DB, no Playwright."""
from __future__ import annotations

import re

_DOMAIN_RE = re.compile(r"^[a-z0-9]([a-z0-9-]*[a-z0-9])?(\.[a-z0-9]([a-z0-9-]*[a-z0-9])?)+(/.*)?$")


def clean_domain(raw: str) -> str:
    """Normalise a domain entry: strip spaces, scheme, leading www., trailing
    slash; lowercase. Keeps any path (some entries target a sub-section)."""
    d = (raw or "").strip().lower()
    for prefix in ("https://", "http://"):
        if d.startswith(prefix):
            d = d[len(prefix):]
    if d.startswith("www."):
        d = d[4:]
    d = d.rstrip("/")
    return d


def dedup_domains(domains) -> list[str]:
    """Clean each entry, drop blanks, dedup while preserving first-seen order."""
    seen = set()
    out = []
    for raw in domains:
        d = clean_domain(raw)
        if not d or d in seen:
            continue
        seen.add(d)
        out.append(d)
    return out


def is_domain_like(d: str) -> bool:
    """True si *d* (déjà passé par clean_domain) ressemble à un host ou
    host/path avec un TLD de 2+ lettres. Rejette les mots seuls, nombres,
    en-têtes de colonnes."""
    if not d or "." not in d:
        return False
    if not _DOMAIN_RE.match(d):
        return False
    host = d.split("/", 1)[0]
    tld = host.rsplit(".", 1)[-1]
    return len(tld) >= 2 and tld.isalpha()


def build_scan_partition(candidates: list[str], existing: set[str]) -> dict:
    """Nettoie + filtre (domain-like) + dédoublonne *candidates*, puis sépare
    en nouveaux (`to_scan`) vs déjà connus (`duplicates`).

    Returns {to_scan, duplicates, invalid_count, total_found} où
    total_found = len(to_scan) + len(duplicates) (valides distincts) et
    invalid_count = nb de candidats rejetés par le filtre domain-like."""
    seen: set[str] = set()
    valid: list[str] = []
    invalid = 0
    for raw in candidates:
        d = clean_domain(raw)
        if not d:
            continue  # cellule vide — pas compté comme invalide
        if not is_domain_like(d):
            invalid += 1
            continue
        if d in seen:
            continue
        seen.add(d)
        valid.append(d)
    existing_clean = {clean_domain(e) for e in existing}
    to_scan = [d for d in valid if d not in existing_clean]
    duplicates = [d for d in valid if d in existing_clean]
    return {
        "to_scan": to_scan,
        "duplicates": duplicates,
        "invalid_count": invalid,
        "total_found": len(valid),
    }


import csv as _csv
import io as _io

_SPLIT_RE = re.compile(r"[\s,;]+")


def extract_from_text(text: str) -> list[str]:
    """Découpe une saisie libre sur retours-ligne, virgules, points-virgules,
    espaces/tabs."""
    if not text:
        return []
    return [p for p in _SPLIT_RE.split(text) if p]


def extract_from_csv(data: bytes) -> list[str]:
    """Aplatit toutes les cellules non vides d'un CSV (UTF-8, fallback latin-1)."""
    try:
        decoded = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        decoded = data.decode("latin-1", errors="ignore")
    out: list[str] = []
    for row in _csv.reader(_io.StringIO(decoded)):
        for cell in row:
            cell = (cell or "").strip()
            if cell:
                out.append(cell)
    return out


def extract_from_xlsx(data: bytes) -> list[str]:
    """Aplatit toutes les cellules non vides d'un classeur .xlsx."""
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell).strip()
                if s:
                    out.append(s)
    wb.close()
    return out


def collect_candidates(text, file_bytes, filename) -> list[str]:
    """Combine la saisie libre et un éventuel fichier (CSV ou XLSX) en une liste
    brute de candidats. Lève ValueError si le fichier n'est ni .csv ni .xlsx."""
    out: list[str] = []
    if text:
        out.extend(extract_from_text(text))
    if file_bytes:
        name = (filename or "").lower()
        if name.endswith(".csv"):
            out.extend(extract_from_csv(file_bytes))
        elif name.endswith(".xlsx"):
            out.extend(extract_from_xlsx(file_bytes))
        else:
            raise ValueError("unsupported_format")
    return out
